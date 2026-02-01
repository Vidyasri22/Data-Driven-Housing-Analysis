import requests
import json
import os
import openpyxl
import warnings
warnings.filterwarnings("ignore")

def excel_data(url, save_path, headers=None):
    try:
        if not headers:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            with open(save_path, 'wb') as f:
                f.write(response.content)
            # print(f"File downloaded successfully at: {save_path}")
            return True
        else:
            print(f"Failed to download file from {url}. Status code: {response.status_code}")
            print(response.headers)  # Print response headers for debugging
    except requests.exceptions.RequestException as e:
        print(f"Error occurred during request: {e}")

def csv_data(url, save_path):
    try:
        response = requests.get(url)
        if response.status_code == 200:
            with open(save_path, 'wb') as f:
                f.write(response.content)
            # print(f"CSV file downloaded successfully to: {save_path}")
        else:
            print(f"Failed to download CSV: HTTP status code {response.status_code}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")


def economic_data(url, save_path,file_name):
    response = requests.get(url)

    if response.status_code == 200:
        data_dict = response.json()

        filename = os.path.join(save_path, file_name)

        with open(filename, 'w') as json_file:
            json.dump(data_dict, json_file)

        # print(f"Data saved to {filename} successfully.")
    else:
        print(f"Failed to fetch data. Status code: {response.status_code}")

def demo_data(url, headers, save_path, file_name):

    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        data_dict = response.json()
        filename = os.path.join(save_path, file_name)
        with open(filename, 'w') as json_file:
            json.dump(data_dict, json_file)

        # print(f"Data saved to {filename} successfully.")
    else:
        print(f"Failed to fetch data. Status code: {response.status_code}")




# Demographic data--------------------------------------------------------------------------------------------------------------------------------------
url_IL = "https://data.ca.gov/api/3/action/datastore_search?resource_id=c8c8bee2-96a6-410f-b32a-2039de52ea12&limit=500"
headers = {
    "accept": "application/json",
    "X-Api-Key": "05237d509fa24fec86aeed72176e59a0"
}
save_IL = r"C:\Users\user1\OneDrive\Documents\University of Southern Calfornia\Spring 24\DSCI 510\project\data\raw\Demographic data"

demo_data(url_IL, headers, save_IL, "IL.json")

url_pph = 'https://dof.ca.gov/wp-content/uploads/sites/352/Forecasting/Demographics/Documents/P4_HHProjections_B2019.xlsx'
save_pph = r'C:\Users\user1\OneDrive\Documents\University of Southern Calfornia\Spring 24\DSCI 510\project\data\raw\Demographic data\PPH.xlsx'

if excel_data(url_pph, save_pph):
    # print("Excel file downloaded successfully!")
    wb = openpyxl.load_workbook(save_pph)
else:
    print("Failed to download the Excel file.")

url_EL = 'https://dof.ca.gov/wp-content/uploads/sites/352/2023/10/2023SeriesW.xlsx'
save_EL = r'C:\Users\user1\OneDrive\Documents\University of Southern Calfornia\Spring 24\DSCI 510\project\data\raw\Demographic data\EL.xlsx'

if excel_data(url_EL, save_EL):
    # print("Excel file downloaded successfully!")
    wb = openpyxl.load_workbook(save_EL)
else:
    print("Failed to download the Excel file.")

# url_births = 'https://dof.ca.gov/wp-content/uploads/sites/352/2023/07/P_CY_Births_Report.xlsx'
# save_births = r'C:\Users\user1\OneDrive\Documents\University of Southern Calfornia\Spring 24\DSCI 510\project\data\raw\Demographic data\births.xlsx'
#
# if excel_data(url_births, save_births):
#     print("Excel file downloaded successfully!")
#     wb = openpyxl.load_workbook(save_births)
# else:
#     print("Failed to download the Excel file.")
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------




# Economic data--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
url_gdp = '''https://apps.bea.gov/api/data/?UserID=689E97AB-F2C8-46B1-BD58-5D356DF2B8C2&method=GetData&datasetname=Regional&TableName=CAGDP1&LineCode=2&Year=2022&GeoFips=CA&ResultFormat=json'''
save_gdp = r'C:\Users\user1\OneDrive\Documents\University of Southern Calfornia\Spring 24\DSCI 510\project\data\raw\Economic data'
economic_data(url_gdp, save_gdp,'GDP.json')

url_PI = '''https://apps.bea.gov/api/data/?UserID=689E97AB-F2C8-46B1-BD58-5D356DF2B8C2&method=GetData&datasetname=Regional&TableName=CAINC1&LineCode=1&Year=2022&GeoFips=CA&ResultFormat=json'''
save_PI = r'C:\Users\user1\OneDrive\Documents\University of Southern Calfornia\Spring 24\DSCI 510\project\data\raw\Economic data'
economic_data(url_PI, save_PI,'PI.json')

url_POP = '''https://apps.bea.gov/api/data/?UserID=689E97AB-F2C8-46B1-BD58-5D356DF2B8C2&method=GetData&datasetname=Regional&TableName=CAINC1&LineCode=2&Year=2022&GeoFips=CA&ResultFormat=json'''
save_POP =  r'C:\Users\user1\OneDrive\Documents\University of Southern Calfornia\Spring 24\DSCI 510\project\data\raw\Economic data'
economic_data(url_POP,save_POP,'POP.json')

url_PCI = '''https://apps.bea.gov/api/data/?UserID=689E97AB-F2C8-46B1-BD58-5D356DF2B8C2&method=GetData&datasetname=Regional&TableName=CAINC1&LineCode=3&Year=2022&GeoFips=CA&ResultFormat=json'''
save_PCI =  r'C:\Users\user1\OneDrive\Documents\University of Southern Calfornia\Spring 24\DSCI 510\project\data\raw\Economic data'
economic_data(url_PCI,save_PCI,'PCI.json')

# url_labourforce = 'https://www.bls.gov/lau/laucnty22.xlsx'
# save_labourforce = r'C:\Users\user1\OneDrive\Documents\University of Southern Calfornia\Spring 24\DSCI 510\project\data\raw\Economic data\labour force.xlsx'
#
# if excel_data(url_labourforce, save_labourforce):
#     print("Excel file downloaded successfully!")
#     wb = openpyxl.load_workbook(save_labourforce)
# else:
#     print("Failed to download the Excel file.")
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------




# Housing data-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# url_min = "https://data.census.gov/table/DECENNIALSF32000.H075?t=Housing:Housing%20Value%20and%20Purchase%20Price&g=010XX00US_040XX00US06$0500000"
# save_min = r"C:\Users\user1\OneDrive\Documents\University of Southern Calfornia\Spring 24\DSCI 510\project\data\raw\Housing data\min_price.csv"
# csv_data(url_min,save_min)
#
# url_med = "https://data.census.gov/table/DECENNIALSF32000.H076?t=Housing:Housing%20Value%20and%20Purchase%20Price&g=010XX00US_040XX00US06$0500000"
# save_med = r"C:\Users\user1\OneDrive\Documents\University of Southern Calfornia\Spring 24\DSCI 510\project\data\raw\Housing data\median_price.csv"
# csv_data(url_med,save_med)
#
# url_max = "https://data.census.gov/table/DECENNIALSF32000.H077?t=Housing:Housing%20Value%20and%20Purchase%20Price&g=010XX00US_040XX00US06$0500000"
# save_max = r"C:\Users\user1\OneDrive\Documents\University of Southern Calfornia\Spring 24\DSCI 510\project\data\raw\Housing data\max_price.csv"
# csv_data(url_max,save_max)
